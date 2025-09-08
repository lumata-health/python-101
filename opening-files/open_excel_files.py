from collections import deque
from pathlib import Path
from openpyxl import load_workbook, Workbook

# REPLACE FILE PATH HERE
path = Path(r'data.xlsx')
N_HEAD = 10
N_TAIL = 10

wb = load_workbook(path, read_only=True, data_only=True)
ws = wb.active  # or wb["Sheet1"] if needed
rows_iter = ws.iter_rows(values_only=True)

header = None
first_rows = []
last_rows = deque(maxlen=N_TAIL)
count = 0

for row in rows_iter:
    # Skip empty rows
    if (row is None) or all((c is None) or (str(c).strip() == "") for c in row):
        continue

    if header is None:
        header = list(row)  # first non-empty row is header
        continue

    count += 1
    if count <= N_HEAD:
        first_rows.append(list(row))
    last_rows.append(list(row))

wb.close()

# Print summary to console
if header is None:
    print("Empty sheet or missing header.")
else:
    print("Columns")
    print(header)

    print(f"\nFirst {min(N_HEAD, count)} rows")
    for r in first_rows:
        print(r)

    print(f"\nLast {min(N_TAIL, count)} rows")
    for r in last_rows:
        print(r)

    print(f"\nTotal data rows (excluding header): {count}")

    # Save small sample workbook: header + head(10) + tail(10)
    sample_path = path.with_name(path.stem + "_SAMPLE.xlsx")

    sample_wb = Workbook()
    sample_ws = sample_wb.active
    sample_ws.title = "sample"

    sample_ws.append(header)
    for r in first_rows:
        sample_ws.append(r)
    for r in last_rows:
        sample_ws.append(r)

    sample_wb.save(sample_path)
    print(f"\nSample workbook written to: {sample_path}")
